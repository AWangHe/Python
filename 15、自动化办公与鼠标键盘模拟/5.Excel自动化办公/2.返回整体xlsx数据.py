# -*- coding: utf-8 -*-
#xlsx  xls
#openpyxl ---> xlsx

from openpyxl.reader.excel import load_workbook

def readXlsx(path):
    dic = {}
    #打开文件
    file = load_workbook(filename=path)
    #所有表格的名称
    #print(file.get_sheet_names())
    sheets = file.get_sheet_names()
    
    for sheetName in sheets:        
        sheet = file.get_sheet_by_name(sheetName)
        #一张表的所有数据
        sheetInfo = []       
        for lineNum in range(1,sheet.max_row + 1):
            lineList = []
            for columnNum in range(1, sheet.max_column + 1):
                #拿数据
                value = sheet.cell(row=lineNum,column=columnNum).value
                lineList.append(value)
            sheetInfo.append(lineList)
        #将一张表的数据存储到字典
        dic[sheetName] = sheetInfo
    return dic
            
                
        
        
#=不能处理xls文件
path = r"F:\Python\data\15、自动化办公与鼠标键盘模拟\5.Excel自动化办公\001.xlsx"
dic = readXlsx(path)
print(dic)














