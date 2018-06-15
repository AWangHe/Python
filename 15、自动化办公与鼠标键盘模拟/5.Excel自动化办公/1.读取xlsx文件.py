# -*- coding: utf-8 -*-
#pip install openpyxl
#xlsx  xls
#openpyxl ---> xlsx

from openpyxl.reader.excel import load_workbook

def readXlsx(path):
    #打开文件
    file = load_workbook(filename=path)
    #所有表格的名称
    #print(file.get_sheet_names())
    sheets = file.get_sheet_names()
    
    #拿出一个表格
    sheet = file.get_sheet_by_name(sheets[0])
    #最大行数
    #print(sheet.max_row)
    #最大列数
    #print(sheet.max_column)
    #表名
    #print(sheet.title)
    for lineNum in range(1,sheet.max_row + 1):
        lineList = []
        for columnNum in range(1, sheet.max_column + 1):
            #拿数据
            value = sheet.cell(row=lineNum,column=columnNum).value
            if value != None:
                lineList.append(value)
        print(lineList)
                
        
        
#=不能处理xls文件
path = r"F:\Python\data\15、自动化办公与鼠标键盘模拟\5.Excel自动化办公\001.xlsx"
readXlsx(path)

















